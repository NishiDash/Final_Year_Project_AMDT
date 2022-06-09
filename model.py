from matplotlib.pyplot import text
import xlsxwriter
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors

def model(file1,attribute):
    try:
        print(Fore.RESET)
        path = './excel files/'+file1
        
        df = pd.read_excel(path)
        n = df.count()[0]+2
        m =  df.count()[2]+3
        p = df.count()[5]+2

        obj = openpyxl.load_workbook(path.strip())
        today = date.today()

        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "output_"+str(today)

        sheet_obj = obj["Sheet1"]
        sheet_obj1 = obj["output_"+str(today)]

        sheet_obj.insert_cols(0,amount=1)
        sheet_obj.insert_cols(4,amount=1)
        sheet_obj.insert_cols(8,amount=1)
        sheet_obj.insert_cols(10,amount=1)

        sheet_obj['A1']='trim id'
        sheet_obj['D1']='trim Turbine_Serial_number'
        sheet_obj['H1']='trim sourceKey'
        sheet_obj['J1']='trim serial number'
        sheet_obj['M1']='r_model(RAMP) wrt id'
        sheet_obj['N1']='model(RAMP) wrt id'
        sheet_obj['O1']='model(Predix) wrt id'
        sheet_obj['P1']='(AWS vs RAMP - R_attribute)'
        sheet_obj['Q1']='(AWS vs RAMP)'
        sheet_obj['R1']='(AWS vs Predix)'
        sheet_obj['S1']='Turbine serial number in AWS?'
        sheet_obj['T1']='Serial number in AWS?'
        print("m,n,p",m,n,p)
        sheet_obj1['A1']='id(AWS)'
        sheet_obj1['B1']='model(AWS)'
        sheet_obj1['C1']='Turbine_Serial_number(RAMP)'
        sheet_obj1['D1']='model(RAMP)'
        sheet_obj1['E1']='serial number(Predix)'
        sheet_obj1['F1']='model(Predix)'
        sheet_obj1['G1']='(AWS vs RAMP)'
        sheet_obj1['H1']='(AWS vs Predix)'

        sheet_obj1['J1']='id(AWS)'
        sheet_obj1['K1']='model(AWS)'
        sheet_obj1['L1']='Turbine_Serial_number(RAMP)'
        sheet_obj1['M1']='r_model(RAMP)'
        sheet_obj1['N1']='(AWS vs RAMP R_Attribute)'
        ##############################FILLING##############################
        f10 = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        for y in range(1,20+1):
            sheet_obj.cell(row=1,column=y).fill = f10
            sheet_obj.cell(row=1,column=y).font = Font(bold=True)
        for y in range(1,8+1):
            sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)
        for y in range(10,14+1):
            sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)
        
        for i in range(2,n):
            i1 = 'A'+str(i)
            i2 = 'M'+str(i)
            i3 = 'N'+str(i)
            i4 = 'O'+str(i)
            i5 = 'P'+str(i)
            i6 = 'Q'+str(i)
            i7 = 'R'+str(i)

            i8 = 'A'+str(i)
            i9 = 'B'+str(i)
            i10 = 'C'+str(i)
            i11 = 'D'+str(i)
            i12 = 'E'+str(i)
            i13 = 'F'+str(i)
            i14 = 'G'+str(i)
            i15 = 'H'+str(i)
            i16 = 'J'+str(i)
            i17 = 'K'+str(i)
            i18 = 'L'+str(i)
            i19 = 'M'+str(i)
            i20 = 'N'+str(i)
            
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=IF(ISNA(VLOOKUP(A'+str(i)+',D:F,3,FALSE)),"AWS id not in RAMP",IF(LEN(VLOOKUP(A'+str(i)+',D:F,3,FALSE))=0,"",VLOOKUP(A'+str(i)+',D:F,3,FALSE)))'
            f3 = '=IF(ISNA(VLOOKUP(A'+str(i)+',D:G,4,FALSE)),"AWS id not in RAMP",IF(LEN(VLOOKUP(A'+str(i)+',D:G,4,FALSE))=0,"",VLOOKUP(A'+str(i)+',D:G,4,FALSE)))'
            f4 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',H:L,5,FALSE)),ISNA(VLOOKUP(A'+str(i)+',J:L,3,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',H:L,5,FALSE))),IF(LEN(VLOOKUP(A'+str(i)+',H:L,5,FALSE))=0,"",VLOOKUP(A'+str(i)+',H:L,5,FALSE)),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',J:L,3,FALSE))),IF(LEN(VLOOKUP(A'+str(i)+',J:L,3,FALSE))=0,"",VLOOKUP(A'+str(i)+',J:L,3,FALSE)),"AWS id not in Predix")))'
            f5 = '=IF(Q'+str(i)+'="matching","",IF(M'+str(i)+'="AWS id not in RAMP","AWS id not in RAMP",IF(AND(OR(C'+str(i)+'="NULL",C'+str(i)+'=""),OR(M'+str(i)+'="NULL",M'+str(i)+'="")),"matching",IF(C'+str(i)+'=M'+str(i)+',"matching","not matching"))))'
            f6 = '=IF(AND(OR(C'+str(i)+'="NULL",C'+str(i)+'=""),OR(N'+str(i)+'="NULL",N'+str(i)+'="")),"matching",IF(N'+str(i)+'="AWS id not in RAMP","AWS id not in RAMP",IF(C'+str(i)+'=N'+str(i)+',"matching","not matching")))'
            f7 = '=IF(O'+str(i)+'="AWS id not in predix","AWS id not in Predix",IF(AND(OR(C'+str(i)+'="NULL",C'+str(i)+'=""),OR(O'+str(i)+'="NULL",O'+str(i)+'="")),"matching",IF(C'+str(i)+'=O'+str(i)+',"matching","not matching")))'

            f8 = '=Sheet1!B'+str(i)
            f9 = '=Sheet1!C'+str(i)
            f11 = '=Sheet1!N'+str(i)
            f13 = '=Sheet1!O'+str(i)
            f14 = '=Sheet1!Q'+str(i)
            f15 = '=Sheet1!R'+str(i)
            f17 = '=Sheet1!C'+str(i)
            f19 = '=Sheet1!M'+str(i)
            f20 = '=Sheet1!P'+str(i)

            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
            sheet_obj[i3]=f3
            sheet_obj[i4]=f4
            sheet_obj[i5]=f5
            sheet_obj[i6]=f6
            sheet_obj[i7]=f7

            sheet_obj1[i8]=f8
            sheet_obj1[i10]=f8
            sheet_obj1[i12]=f8
            sheet_obj1[i16]=f8
            sheet_obj1[i18]=f8
            sheet_obj1[i9]=f9
            sheet_obj1[i11]=f11
            sheet_obj1[i13]=f13
            sheet_obj1[i14]=f14
            sheet_obj1[i15]=f15
            sheet_obj1[i17]=f17
            sheet_obj1[i19]=f19
            sheet_obj1[i20]=f20

        for i in range(2,m):
            i1 = 'D'+str(i)
            i2 = 'S'+str(i)
            f1 = '=TRIM(E'+str(i)+')'
            f2 = '=IF(ISNA(VLOOKUP(D'+str(i)+',A:A,1,FALSE)),"RAMP id not in AWS",VLOOKUP(D'+str(i)+',A:A,1,FALSE))'
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
        for i in range(2,p):
            i1 = 'H'+str(i)
            i2 = 'J'+str(i)
            i3 = 'T'+str(i)
            f1 = '=TRIM(I'+str(i)+')'
            f2 = '=TRIM(K'+str(i)+')'
            f3  = '=IF(AND(ISNA(VLOOKUP(H'+str(i)+',A:A,1,FALSE)),ISNA(VLOOKUP(J'+str(i)+',A:A,1,FALSE))),"Predix id not in AWS",IF(NOT(ISNA(VLOOKUP(H'+str(i)+',A:A,1,FALSE))),VLOOKUP(H'+str(i)+',A:A,1,FALSE),IF(NOT(ISNA(VLOOKUP(J'+str(i)+',A:A,1,FALSE))),VLOOKUP(J'+str(i)+',A:A,1,FALSE),"Predix id not in AWS")))'
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
            sheet_obj[i3]=f3
        
        i=2
        for j in range(n,m+n):
            ind1 = 'C'+str(j)
            ind2 = 'D'+str(j)
            ind3 = 'G'+str(j)

            frm1 = '=IF(Sheet1!S'+str(i)+'="RAMP id not in AWS",Sheet1!D'+str(i)+',"")'
            frm2 = '=if(Sheet1!S'+str(i)+'="RAMP id not in AWS","RAMP id not in AWS","")'
            frm3 = '=if(Sheet1!S'+str(i)+'="RAMP id not in AWS","RAMP id not in AWS","")'
            
            sheet_obj1[ind1]=frm1
            sheet_obj1[ind2]=frm2
            sheet_obj1[ind3]=frm3
            i=i+1
        i=2
        for j in range(n,p+n):
            ind1 = 'E'+str(j)
            ind2 = 'F'+str(j)
            ind3 = 'H'+str(j)

            frm1 = '=if(Sheet1!T'+str(i)+'="Predix id not in AWS",Sheet1!I'+str(i)+',"")'
            frm2 = '=if(Sheet1!T'+str(i)+'="Predix id not in AWS","Predix id not in AWS","")'
            frm3 = '=if(Sheet1!T'+str(i)+'="Predix id not in AWS","Predix id not in AWS","")'
            
            sheet_obj1[ind1]=frm1
            sheet_obj1[ind2]=frm2
            sheet_obj1[ind3]=frm3
            i=i+1
        obj.save(path)

    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")