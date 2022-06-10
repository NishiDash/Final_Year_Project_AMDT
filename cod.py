from matplotlib.pyplot import text
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles import Font

def cod(file1,attribute):
    try:
        print(Fore.RESET)
        path = './excel files/'+file1
        df = pd.read_excel(path,sheet_name="Sheet1")
        m = df.count()[2]+3
        n = df.count()[0]+2
        
        obj = openpyxl.load_workbook(path.strip())
        today = date.today()

        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "output_"+str(today)+"_COD"
        print('n: ',n," m: ",m)
        sheet_obj = obj["Sheet1"]
        sheet_obj1 = obj["output_"+str(today)+"_COD"]

        sheet_obj.insert_cols(0,amount=1)
        sheet_obj.insert_cols(4,amount=1)
        sheet_obj.insert_cols(5,amount=1)
        sheet_obj.insert_cols(8,amount=1)
        # sheet_obj.insert_cols(9,amount=1) 

        
        sheet_obj['A1']='TRIM_id'
        sheet_obj['D1']='Appropriate format of '+attribute
        sheet_obj['E1']='Trim_Turbine Serial Number'
        sheet_obj['K1']='R_'+attribute+' wrt id'
        sheet_obj['L1']=attribute+' wrt id'
        sheet_obj['H1']='Appropriate format of R_'+attribute+' wrt id'
        sheet_obj['J1']='Appropriate format of '+attribute+' wrt id'
        sheet_obj['M1']= attribute+'_Discrepancy'
        sheet_obj['N1']= 'R_'+attribute+'_Discrepancy'
        sheet_obj['O1']='ID in AWS?'
        
        print("m,n",m,n)
        sheet_obj1['A1']='Id'
        sheet_obj1['B1']=attribute+" (AWS)"
        sheet_obj1['C1']='Turbine Serial Number'
        sheet_obj1['D1']='R_'+attribute+" (RAMP)"
        sheet_obj1['E1']='R_'+attribute+'_Discrepancy'

        sheet_obj1['G1']='Id'
        sheet_obj1['H1']=attribute+" (AWS)"
        sheet_obj1['I1']='Turbine Serial Number'
        sheet_obj1['J1']=attribute+" (RAMP)"
        sheet_obj1['K1']=attribute+'_Discrepancy'
        ##############################FILLING##############################
        f10 = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        for y in range(1,15+1):
            sheet_obj.cell(row=1,column=y).fill = f10
            sheet_obj.cell(row=1,column=y).font = Font(bold=True)
        for y in range(1,5+1):
            sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)
        for y in range(7,11+1):
            sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)
            
        for j in range(2,n):
            index1 = 'A'+str(j)
            index2 = 'D'+str(j)
            
            formula1 = '=TRIM(B'+str(j)+')'
            formula2 = '=IF(C'+str(j)+'="NULL","",IF(C'+str(j)+'="","",IF(ISERROR(TEXT(DATE(YEAR(C'+str(j)+'),MONTH(C'+str(j)+'),DAY(C'+str(j)+')),"m/d/yyyy")),C'+str(j)+',TEXT(DATE(YEAR(C'+str(j)+'),MONTH(C'+str(j)+'),DAY(C'+str(j)+')),"m/d/yyyy"))))'
            
            sheet_obj[index1]= formula1
            sheet_obj[index2]= formula2

        for j in range(2,m):
            index1 = 'E'+str(j)
            index2 = 'O'+str(j)

            formula1 = '=TRIM(F'+str(j)+')'
            formula2 = '=IF(ISNA(VLOOKUP(E'+str(j)+',A:A,1,FALSE)),"RAMP id not in AWS",IF( LEN(VLOOKUP(E'+str(j)+',A:A,1,FALSE))=0,"",VLOOKUP(E'+str(j)+',A:A,1,FALSE)))' 

            sheet_obj[index1]=formula1
            sheet_obj[index2]=formula2    
            
        for j in range(2,n):
            index1 = 'K'+str(j)
            index2 = 'L'+str(j)
            index3 = 'H'+str(j)
            index4 = 'J'+str(j)

            formula1 = '=IF(ISNA(VLOOKUP(A'+str(j)+',E:J,6,FALSE)),"AWS id not in RAMP",IF(LEN(VLOOKUP(A'+str(j)+',E:J,6,FALSE))=0,"",VLOOKUP(A'+str(j)+',E:J,6,FALSE)))'
            formula2 = '=IF(ISNA(VLOOKUP(A'+str(j)+',E:H,4,FALSE)),"AWS id not in RAMP",IF(LEN(VLOOKUP(A'+str(j)+',E:H,4,FALSE))=0,"",VLOOKUP(A'+str(j)+',E:H,4,FALSE)))'
            formula3 = '=IF(G'+str(j)+'="NULL","",IF(G'+str(j)+'="","",IF(G'+str(j)+'="AWS id not in RAMP","AWS id not in RAMP",TEXT(DATE(YEAR(G'+str(j)+'),MONTH(G'+str(j)+'),DAY(G'+str(j)+')),"m/d/yyyy"))))'
            formula4 = '=IF(I'+str(j)+'="NULL","",IF(I'+str(j)+'="","",IF(I'+str(j)+'="AWS id not in RAMP","AWS id not in RAMP",TEXT(DATE(YEAR(I'+str(j)+'),MONTH(I'+str(j)+'),DAY(I'+str(j)+')),"m/d/yyyy"))))'

            sheet_obj[index1]=formula1
            sheet_obj[index2]=formula2  
            sheet_obj[index3]=formula3
            sheet_obj[index4]=formula4

        for j in range(2,n):
            index1 = 'M'+str(j)
            index2 = 'N'+str(j)
            formula1 = '=IF(L'+str(j)+'="AWS id not in RAMP","AWS id not in RAMP",IF(AND(OR(D'+str(j)+'="",D'+str(j)+'="NULL"),OR(L'+str(j)+'="",L'+str(j)+'="NULL")),"matching",IF(D'+str(j)+'=L'+str(j)+',"matching","not matching")))'
            formula2 = '=IF(K'+str(j)+'="AWS id not in RAMP","AWS id not in RAMP",IF(M'+str(j)+'="matching","",IF(AND(OR(D'+str(j)+'="",D'+str(j)+'="NULL"),OR(K'+str(j)+'="",K'+str(j)+'="NULL")),"matching",IF(D'+str(j)+'=K'+str(j)+',"matching","not matching"))))'
            sheet_obj[index1]=formula1
            sheet_obj[index2]=formula2 
              
        for j in range(2,n):
            ind1 = 'A'+str(j)
            ind2 = 'B'+str(j)
            ind3 = 'C'+str(j)
            ind4 = 'D'+str(j)
            ind5 = 'E'+str(j)      

            frm1 = '=Sheet1!B'+str(j)
            frm2 = '=Sheet1!D'+str(j)
            frm3 = '=Sheet1!B'+str(j)
            frm4 = '=Sheet1!L'+str(j)
            frm5 = '=Sheet1!M'+str(j)

            sheet_obj1[ind1] = frm1
            sheet_obj1[ind2] = frm2
            sheet_obj1[ind3] = frm3
            sheet_obj1[ind4] = frm4
            sheet_obj1[ind5] = frm5

            ind1 = 'G'+str(j)
            ind2 = 'H'+str(j)
            ind3 = 'I'+str(j)
            ind4 = 'J'+str(j)
            ind5 = 'K'+str(j) 

            frm1 = '=Sheet1!B'+str(j)
            frm2 = '=Sheet1!D'+str(j)
            frm3 = '=Sheet1!B'+str(j)
            frm4 = '=Sheet1!K'+str(j)
            frm5 = '=Sheet1!N'+str(j)            

            sheet_obj1[ind1] = frm1
            sheet_obj1[ind2] = frm2
            sheet_obj1[ind3] = frm3
            sheet_obj1[ind4] = frm4
            sheet_obj1[ind5] = frm5
            
        
        i1 = 'C'+str(n)
        i2 = 'E'+str(n)
        f1 = 'FILTER(Sheet1!F2:F'+str(m-1)+',Sheet1!O2:O'+str(m-1)+'="RAMP id not in AWS")'
        f2 = 'FILTER(Sheet1!O2:O'+str(m-1)+',Sheet1!O2:O'+str(m-1)+'="RAMP id not in AWS")'
        sheet_obj1[i1]=f1
        sheet_obj1[i2]=f2

        obj.save(path)
        
            
    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")