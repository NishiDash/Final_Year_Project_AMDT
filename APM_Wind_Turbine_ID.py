from matplotlib.pyplot import text
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles import Font

def APM(file1):
    try:
        print(Fore.RESET)
        path = './excel files/'+file1
        
        df = pd.read_excel(path)
        n = df.count()[0]+2
        m =  df.count()[2]+3
        # p = df.count()[4]+2

        obj = openpyxl.load_workbook(path.strip())
        today = date.today()

        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "output_"+str(today)

        sheet_obj = obj["Sheet1"]
        sheet_obj1 = obj["output_"+str(today)]

        sheet_obj.insert_cols(0,amount=1)
        sheet_obj.insert_cols(4,amount=1)
        sheet_obj.insert_cols(6,amount=1)
        # sheet_obj.insert_cols(9,amount=1)
        sheet_obj['A1']='trim id'
        sheet_obj['D1']='trim SourceKey'
        sheet_obj['F1']='trim Serial Number'
        sheet_obj['I1']='APM(Predix) wrt id'
        sheet_obj['J1']='(AWS vs Predix)'
        sheet_obj['K1']='Serial number in AWS?'
        
        print("m,n",m,n)
        sheet_obj1['A1']='id(AWS)'
        sheet_obj1['B1']='apm wind turbine ID(AWS)'
        sheet_obj1['C1']='serial number(Predix)'
        sheet_obj1['D1']='apm wind turbine ID(Predix)'
        sheet_obj1['E1']='(AWS vs Predix)'
        ##############################FILLING##############################
        
        # from openpyxl.styles import Font
        for y in range(1,11+1):
            sheet_obj.cell(row=1,column=y).fill = f1
            sheet_obj.cell(row=1,column=y).font = Font(bold=True)
        for y in range(1,5+1):
            sheet_obj1.cell(row=1,column=y).fill = f1
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)
        for i in range(2,n):
            i1 = 'A'+str(i)
            i2 = 'I'+str(i)
            i3 = 'J'+str(i)
            i4 = 'A'+str(i)
            i5 = 'C'+str(i)
            i6 = 'B'+str(i)
            i7 = 'D'+str(i)
            i8 = 'E'+str(i)
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',D:H,5,FALSE)),ISNA(VLOOKUP(A'+str(i)+',F:H,3,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',D:H,5,FALSE))),IF(LEN(VLOOKUP(A'+str(i)+',D:H,5,FALSE))=0,"",TRIM(VLOOKUP(A'+str(i)+',D:H,5,FALSE))),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',F:H,3,FALSE))),IF(LEN(VLOOKUP(A'+str(i)+',F:H,3,FALSE))=0,"",TRIM(VLOOKUP(A'+str(i)+',F:H,3,FALSE))))))'
            f3 = '=IF(I'+str(i)+'="AWS id not in Predix","AWS id not in Predix",IF(C'+str(i)+'="Id not in alias","Id not in alias",IF(AND(OR(C'+str(i)+'="",C'+str(i)+'="NULL"),OR(I'+str(i)+'="",I'+str(i)+'="NULL")),"matching",IF(C'+str(i)+'=I'+str(i)+',"matching","not matching"))))'
            f4 = '=Sheet1!B'+str(i)
            f5 = '=IF(Sheet1!C'+str(i)+'="Id not in alias","",Sheet1!C'+str(i)+')'
            f6 = '=Sheet1!I'+str(i)
            f7 = '=Sheet1!J'+str(i)
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
            sheet_obj[i3]=f3
            sheet_obj1[i4]=f4
            sheet_obj1[i5]=f4
            sheet_obj1[i6]=f5
            sheet_obj1[i7]=f6
            sheet_obj1[i8]=f7
     
        for i in range(2,m):
            i1 = 'D'+str(i)
            i2 = 'F'+str(i)
            i3 = 'K'+str(i)
            f1 = '=TRIM(E'+str(i)+')'
            f2 = '=TRIM(G'+str(i)+')'
            f3 = '=IF(AND(ISNA(VLOOKUP(D'+str(i)+',A:A,1,FALSE)),ISNA(VLOOKUP(F'+str(i)+',A:A,1,FALSE))),"Predix id not in AWS",IF(NOT(ISNA(VLOOKUP(D'+str(i)+',A:A,1,FALSE))),VLOOKUP(D'+str(i)+',A:A,1,FALSE),IF(NOT(ISNA(VLOOKUP(F'+str(i)+',A:A,1,FALSE))),VLOOKUP(F'+str(i)+',A:A,1,FALSE))))'
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
            sheet_obj[i3]=f3
        i=2
        for j in range(n,m+n):
            ind1 = 'C'+str(j)
            ind2 = 'D'+str(j)
            ind3 = 'E'+str(j)

            frm1 = '=IF(Sheet1!K'+str(i)+'="Predix id not in AWS",Sheet1!D'+str(i)+',"")'
            frm2 = '=if(Sheet1!K'+str(i)+'="Predix id not in AWS","Predix id not in AWS","")'
            frm3 = '=if(Sheet1!K'+str(i)+'="Predix id not in AWS","Predix id not in AWS","")'
            
            sheet_obj1[ind1]=frm1
            sheet_obj1[ind2]=frm2
            sheet_obj1[ind3]=frm3
            i=i+1

        obj.save(path)
    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")