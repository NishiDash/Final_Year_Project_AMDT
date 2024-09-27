from matplotlib.pyplot import text
import xlsxwriter
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles import Font
def make(file1,attribute):
    try:
        print(Fore.RESET)
        path = './excel files/'+file1
        df = pd.read_excel(path,sheet_name="Sheet1")
        n = df.count()[2] + 3 
        m = df.count()[0] + 2
        p = df.count()[5] + 2
        
        obj = openpyxl.load_workbook(path.strip())
        today = date.today()

        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "output_"+str(today)
        
        sheet_obj = obj["Sheet1"]
        sheet_obj1 = obj["output_"+str(today)]

        sheet_obj['G1']='OEM Supplier'
        sheet_obj['H1']='Trim_id'
        sheet_obj['I1']='OEM Supplier wrt id'
        sheet_obj['J1']='OEM Supplier Discrepancy'
        sheet_obj['K1']='Turbine Serial Number'
        sheet_obj['L1']='ID Present in AWS?'

        sheet_obj1['A1']='Id'
        sheet_obj1['B1']=attribute+" (AWS)"
        sheet_obj1['C1']='Turbine Serial Number'
        sheet_obj1['D1']=attribute+" (RAMP)"
        sheet_obj1['E1']=attribute+"_Discrepancy"

        ##############################FILLING##############################
        f10 = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        for y in range(1,12+1):
            # sheet_obj.cell(row=1,column=y).fill = f10
            sheet_obj.cell(row=1,column=y).font = Font(bold=True)
        for y in range(1,5+1):
            # sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)

        for i in range(2,n):
            index='G'+str(i)
            formula = '=VLOOKUP(C'+str(i)+',D:E,2,FALSE)' 
            sheet_obj[index]= formula  
        for j in range(2,m):
            index1 = 'H'+str(j)
            index2 = 'I'+str(j)
            index3 = 'J'+str(j)
            
            formula1 = '=TRIM(A'+str(j)+')'
            formula2 = '=if(ISNA(VLOOKUP(H'+str(j)+',F:G,2,FALSE)),"AWS idd not in RAMP",if(len(VLOOKUP(H'+str(j)+',F:G,2,FALSE))=0,"",VLOOKUP(H'+str(j)+',F:G,2,FALSE)))'
            formula3 = '=if(I'+str(j)+'="AWS id not in RAMP","AWS id not in RAMP",if(AND(OR(B'+str(j)+'="NULL",B'+str(j)+'=""),OR(I'+str(j)+'="NULL",I'+str(j)+'="")),"matching",if(B'+str(j)+'<>I'+str(j)+',"not matching","matching")) )'
            
            sheet_obj[index1]= formula1
            sheet_obj[index2]= formula2
            sheet_obj[index3]= formula3
        for j in range(2,m):
            ind1 = 'A'+str(j)
            ind2 = 'B'+str(j)
            ind3 = 'C'+str(j)
            ind4 = 'D'+str(j)
            ind5 = 'E'+str(j)

            frm1 = '=Sheet1!A'+str(j)
            frm2 = '=Sheet1!B'+str(j)
            frm3 = '=Sheet1!A'+str(j)
            frm4 = '=Sheet1!I'+str(j)
            frm5 = '=Sheet1!J'+str(j)

            sheet_obj1[ind1] = frm1
            sheet_obj1[ind2] = frm2
            sheet_obj1[ind3] = frm3
            sheet_obj1[ind4] = frm4
            sheet_obj1[ind5] = frm5
        for j in range(2,p+1):
            index1 = 'K'+str(j)
            index2 = 'L'+str(j)

            formula1 = '=TRIM(F'+str(j)+')'
            formula2 = '=if(ISNA(vlookup(K'+str(j)+',A:B,2,false)),"RAMP id not in AWS",if( len(vlookup(K'+str(j)+',A:B,2,false))=0,"",vlookup(K'+str(j)+',A:B,2,false) ))'

            sheet_obj[index1]=formula1
            sheet_obj[index2]=formula2
        i1 = 'C'+str(m)
        i2 = 'E'+str(m)
        f1 = 'FILTER(Sheet1!D2:D'+str(p-1)+',Sheet1!L2:L'+str(p-1)+'="RAMP id not in AWS")'
        f2 = 'FILTER(Sheet1!L2:L'+str(p-1)+',Sheet1!L2:L'+str(p-1)+'="RAMP id not in AWS")'
        sheet_obj1[i1]=f1
        sheet_obj1[i2]=f2
    
        obj.save(path)
    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")