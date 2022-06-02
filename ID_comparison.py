import keyword
from matplotlib.pyplot import text
import xlsxwriter
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
import subprocess
import os
import time
import numpy as np
from pynput.keyboard import Key,Controller
def ID_comp(file1):
    try:
        print(Fore.RESET)
        path = './excel files/'+file1
        print(path)

        df1 = pd.read_excel(path,sheet_name="Sheet1")
        n = df1.count()[0] + 2
        m = df1.count()[1] + 3
        p = df1.count()[2] + 2
       
        print("n:",n," m:",m," p:",p)
        today = date.today()
        obj = openpyxl.load_workbook(path.strip())
        
        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "ID Comparison"


        sheet_obj1 = obj["Sheet1"]
        sheet_obj2 = obj["ID Comparison"]

        sheet_obj1.insert_cols(0,amount=1)
        sheet_obj1.insert_cols(3,amount=1)
        sheet_obj1.insert_cols(5,amount=1)
        sheet_obj1.insert_cols(7,amount=1)
        Clr_fill = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        sheet_obj1['A1']='trim id'
        sheet_obj1['C1']='trim Turbine_Serial_Number'
        sheet_obj1['E1']='trim Source Key'
        sheet_obj1['G1']='trim Serial Number'
        sheet_obj1['I1']='AWS ID not in RAMP'
        sheet_obj1['J1']='AWS ID not in Predix'
        sheet_obj1['K1']='RAMP ID not in AWS'
        sheet_obj1['L1']='RAMP ID not in Predix'
        sheet_obj1['M1']='Predix ID not in AWS'
        sheet_obj1['N1']='Predix ID not in RAMP'
    
        sheet_obj2['A2']="AWS ID not in RAMP"
        sheet_obj2['B2']="AWS ID not in Predix"
        sheet_obj2['A2'].fill = Clr_fill
        sheet_obj2['A2'].font = Font(bold=True)
        sheet_obj2['B2'].fill = Clr_fill
        sheet_obj2['B2'].font = Font(bold=True)

        sheet_obj2['C2']="RAMP ID not in AWS"
        sheet_obj2['D2']="RAMP ID not in Predix"
        sheet_obj2['C2'].fill = Clr_fill
        sheet_obj2['C2'].font = Font(bold=True)
        sheet_obj2['D2'].fill = Clr_fill
        sheet_obj2['D2'].font = Font(bold=True)

        sheet_obj2['E2']="Predix ID not in AWS"
        sheet_obj2['F2']="Predix ID not in RAMP"
        sheet_obj2['E2'].fill = Clr_fill
        sheet_obj2['E2'].font = Font(bold=True)
        sheet_obj2['F2'].fill = Clr_fill
        sheet_obj2['F2'].font = Font(bold=True)
        ##############################FILLING##############################
        f10 = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        for y in range(1,14+1):
            sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)

        for i in range(2,n):
            i1 = 'A'+str(i)
            i2 = 'I'+str(i)
            i3 = 'J'+str(i)
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=IF(ISNA(VLOOKUP(A'+str(i)+',C:D,1,FALSE)),"AWS id not in RAMP",VLOOKUP(A'+str(i)+',C:D,1,FALSE))'
            f3 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',E:F,1,FALSE)),ISNA(VLOOKUP(A'+str(i)+',G:H,1,FALSE))), "AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',E:F,1,FALSE))),VLOOKUP(A'+str(i)+',E:F,1,FALSE),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',G:H,1,FALSE))),VLOOKUP(A'+str(i)+',G:H,1,FALSE),"AWS id not in Predix")))'
            sheet_obj1[i1]=f1
            sheet_obj1[i2]=f2
            sheet_obj1[i3]=f3
        for i in range(2,m):
            i1 = 'C'+str(i)
            i2 = 'K'+str(i)
            i3 = 'L'+str(i)
            f1 = '=TRIM(D'+str(i)+')'
            f2 = '=IF(ISNA(VLOOKUP(C'+str(i)+',A:B,1,FALSE)),"RAMP id not in AWS",VLOOKUP(C'+str(i)+',A:B,1,FALSE))'
            f3 = '=IF(AND(ISNA(VLOOKUP(C'+str(i)+',E:E,1,FALSE)),ISNA(VLOOKUP(C'+str(i)+',G:G,1,FALSE))),"RAMP id not in Predix",IF(NOT(ISNA(VLOOKUP(C'+str(i)+',E:E,1,FALSE))),VLOOKUP(C'+str(i)+',E:E,1,FALSE),IF(NOT(ISNA(VLOOKUP(C'+str(i)+',G:G,1,FALSE))),VLOOKUP(C'+str(i)+',G:G,1,FALSE),"RAMP id not in Predix")))'
            sheet_obj1[i1]=f1
            sheet_obj1[i2]=f2
            sheet_obj1[i3]=f3
        for i in range(2,p):
            i1 = 'E'+str(i)
            i2 = 'G'+str(i)
            i3 = 'M'+str(i)
            i4 = 'N'+str(i)
            f1 = '=TRIM(F'+str(i)+')'
            f2 = '=TRIM(H'+str(i)+')'
            f3 = '=IF(AND(ISNA(VLOOKUP(E'+str(i)+',A:A,1,FALSE)),ISNA(VLOOKUP(G'+str(i)+',A:A,1,FALSE))),"Predix id not in AWS",IF(NOT(ISNA(VLOOKUP(E'+str(i)+',A:A,1,FALSE))),VLOOKUP(E'+str(i)+',A:A,1,FALSE),IF(NOT(ISNA(VLOOKUP(G'+str(i)+',A:A,1,FALSE))),VLOOKUP(G'+str(i)+',A:A,1,FALSE),"Predix id not in AWS")))'
            f4 = '=IF(AND(ISNA(VLOOKUP(E'+str(i)+',C:C,1,FALSE)),ISNA(VLOOKUP(G'+str(i)+',C:C,1,FALSE))),"Predix id not in RAMP",IF(NOT(ISNA(VLOOKUP(E'+str(i)+',C:C,1,FALSE))),VLOOKUP(E'+str(i)+',C:C,1,FALSE),IF(NOT(ISNA(VLOOKUP(G'+str(i)+',C:C,1,FALSE))),VLOOKUP(G'+str(i)+',C:C,1,FALSE),"Predix id not in RAMP")))'
            sheet_obj1[i1]=f1
            sheet_obj1[i2]=f2
            sheet_obj1[i3]=f3
            sheet_obj1[i4]=f4
        obj.save(path)
        print("opening...")
        excel_prog =r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE'
        time.sleep(5)
        subprocess.Popen([excel_prog, os.path.join(os.getcwd(), path)])

        time.sleep(25)
        keyboard = Controller()
        with keyboard.pressed(Key.ctrl):
            keyboard.press('s')
            keyboard.release('s')
        print("Saving...")
        time.sleep(10)
        with keyboard.pressed(Key.alt):
            keyboard.press(Key.f4)
            keyboard.release(Key.f4)
        print("closing...")
        time.sleep(10)
        obj = openpyxl.load_workbook(path, data_only=True)
        df2 = pd.read_excel(path,sheet_name="Sheet1") 

        sheet_obj2 = obj["ID Comparison"]
        

        a = df2.iloc[:,8]
        b = df2.iloc[:,9]
        l1 = list(np.where(a == "AWS id not in RAMP")[0])
        l2 = list(np.where(b == "AWS id not in Predix")[0])
        L1 = len(l1)
        L2 = len(l2)

        print("L1:",L1," L2:",L2)

        sheet_obj2['A1'] = str(L1)
        sheet_obj2['B1'] = str(L2)

        j = 3
        for i in range(0,L1):
            i1 = 'A'+str(j)
            sheet_obj2[i1] = df2[df2.columns[1]][l1[i]]

            j=j+1
        j=3
        for i in range(0,L2):
            i1 = 'B'+str(j)
            sheet_obj2[i1] = df2[df2.columns[1]][l2[i]]
            j=j+1
        print('done')
        #########################################
        a = df2.iloc[:,10]
        b = df2.iloc[:,11]
        l1 = list(np.where(a == "RAMP id not in AWS")[0])
        l2 = list(np.where(b == "RAMP id not in Predix")[0])
        L1 = len(l1)
        L2 = len(l2)
        sheet_obj2['C1'] = str(L1)
        sheet_obj2['D1'] =str(L2)
        print("L1:",L1," L2:",L2)
        j = 3
        for i in range(0,L1):
            i1 = 'C'+str(j)
            sheet_obj2[i1] = df2[df2.columns[3]][l1[i]]
            j=j+1
        j=3
        for i in range(0,L2):
            i1 = 'D'+str(j)
            sheet_obj2[i1] = df2[df2.columns[3]][l2[i]]
            j=j+1
        print('done')
        # #########################################
        a = df2.iloc[:,12]
        b = df2.iloc[:,13]
        l1 = list(np.where(a == "Predix id not in AWS")[0])
        l2 = list(np.where(b == "Predix id not in RAMP")[0])
        L1 = len(l1)
        L2 = len(l2)
        sheet_obj2['E1'] =str(L1)
        sheet_obj2['F1'] =str(L2)
        print("L1:",L1," L2:",L2)
        j = 3
        for i in range(0,L1):
            i1 = 'E'+str(j)
            sheet_obj2[i1].value = df2[df2.columns[5]][l1[i]]
            j=j+1
        j=3
        for i in range(0,L2):
            i1 = 'F'+str(j)
            sheet_obj2[i1].value = df2[df2.columns[5]][l2[i]]
            j=j+1
        print('done')
        obj.save(path)

    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("Error : An error has occured pls verify")
    print(Fore.GREEN + "###################### Successfully! Excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")