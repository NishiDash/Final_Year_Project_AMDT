from matplotlib.pyplot import text
import xlsxwriter
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors

def altitude(file1):
    try: 
        print(Fore.RESET)
        path = './excel files/'+file1
        
        df = pd.read_excel(path,sheet_name="Sheet1")
        n = df.count()[0]+2
        m =  df.count()[2]+3
        p = df.count()[4]+2

        obj = openpyxl.load_workbook(path.strip())
        today = date.today()

        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "output_"+str(today)

        sheet_obj = obj["Sheet1"]
        sheet_obj1 = obj["output_"+str(today)]

        sheet_obj.insert_cols(0,amount=1)
        sheet_obj.insert_cols(4,amount=1)
        sheet_obj.insert_cols(7,amount=1)
        sheet_obj.insert_cols(9,amount=1)

        sheet_obj['A1']='trim id'
        sheet_obj['D1']='trim Turbine_Serial_number'
        sheet_obj['G1']='trim sourceKey'
        sheet_obj['I1']='trim serial number'
        sheet_obj['L1']='altitude(RAMP) wrt id'
        sheet_obj['M1']='altitude(Predix) wrt id'
        sheet_obj['N1']='(AWS vs RAMP)'
        sheet_obj['O1']='(AWS vs Predix)'
        sheet_obj['P1']='Turbine serial number in AWS?'
        sheet_obj['Q1']='Serial number in AWS?'
        
        print("m,n,p",m,n,p)
        sheet_obj1['A1']='id(AWS)'
        sheet_obj1['B1']='hub_height(AWS)'
        sheet_obj1['C1']='Turbine_Serial_number(RAMP)'
        sheet_obj1['D1']='hub_height(RAMP)'
        sheet_obj1['E1']='serial number(Predix)'
        sheet_obj1['F1']='hub_height(Predix)'
        sheet_obj1['G1']='(AWS vs RAMP)'
        sheet_obj1['H1']='(AWS vs Predix)'
        ##############################FILLING##############################
        f10 = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        for y in range(1,17+1):
            # sheet_obj.cell(row=1,column=y).fill = f10
            sheet_obj.cell(row=1,column=y).font = Font(bold=True)
        for y in range(1,8+1):
            # sheet_obj1.cell(row=1,column=y).fill = f10
            sheet_obj1.cell(row=1,column=y).font = Font(bold=True)
        for i in range(2,n):
            i1 = 'A'+str(i)
            i2 = 'L'+str(i)
            i3 = 'M'+str(i)
            i4 = 'N'+str(i)
            i5 = 'O'+str(i)
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=IF(ISNA(VLOOKUP(A'+str(i)+',D:F,3,FALSE)),"AWS id not in RAMP",IF(LEN(VLOOKUP(A'+str(i)+',D:F,3,FALSE))=0,"",VLOOKUP(A'+str(i)+',D:F,3,FALSE)))'
            f3 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',G:K,5,FALSE)),ISNA(VLOOKUP(A'+str(i)+',I:K,3,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',G:K,5,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',G:K,5,FALSE))=0,VLOOKUP(A'+str(i)+',G:K,5,FALSE)="NA"),"",TRIM(VLOOKUP(A'+str(i)+',G:K,5,FALSE))),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',I:K,3,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',I:K,3,FALSE))=0,VLOOKUP(A'+str(i)+',I:K,3,FALSE)="NA"),"",TRIM(VLOOKUP(A'+str(i)+',I:K,3,FALSE))),"AWS id not Predix")))'
            f4 = '=IF(L'+str(i)+'="AWS id not in RAMP","AWS id not in RAMP",IF(AND(OR(C'+str(i)+'="NULL",C'+str(i)+'=0,C'+str(i)+'=""),OR(L'+str(i)+'="NULL",L'+str(i)+'=0,L'+str(i)+'="")),"matching",IF(C'+str(i)+'=L'+str(i)+',"matching","not matching")))'
            f5 = '=IF(M'+str(i)+'="AWS id not in Predix","AWS id not in Predix",IF(AND(OR(C'+str(i)+'="NULL",C'+str(i)+'=0,C'+str(i)+'=""),OR(M'+str(i)+'="NULL",M'+str(i)+'=0,M'+str(i)+'="")),"matching",IF(TRIM(C'+str(i)+')=M'+str(i)+',"matching","not matching")))'
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
            sheet_obj[i3]=f3
            sheet_obj[i4]=f4
            sheet_obj[i5]=f5
        for i in range(2,m):
            i1 = 'D'+str(i)
            i2 = 'P'+str(i)
            f1 = '=TRIM(E'+str(i)+')'
            f2 = '=IF(ISNA(VLOOKUP(D'+str(i)+',A:B,1,FALSE)),"RAMP id not in AWS",VLOOKUP(D'+str(i)+',A:B,1,FALSE))'
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
        for i in range(2,p):
            i1 = 'G'+str(i)
            i2 = 'I'+str(i)
            i3 = 'Q'+str(i)
            f1 = '=TRIM(H'+str(i)+')'
            f2 = '=TRIM(J'+str(i)+')'
            f3 = '=IF(AND(ISNA(VLOOKUP(G'+str(i)+',A:B,1,FALSE)),ISNA(VLOOKUP(I'+str(i)+',A:B,1,FALSE))),"Predix id not in AWS",IF(NOT(ISNA(VLOOKUP(G'+str(i)+',A:B,1,FALSE))),VLOOKUP(G'+str(i)+',A:B,1,FALSE),IF(NOT(ISNA(VLOOKUP(I'+str(i)+',A:B,1,FALSE))),VLOOKUP(I'+str(i)+',A:B,1,FALSE),"Predix id not in AWS")))'
            sheet_obj[i1]=f1
            sheet_obj[i2]=f2
            sheet_obj[i3]=f3
        for i in range(2,n):
            i1 = 'A'+str(i)
            i2 = 'B'+str(i)
            i3 = 'C'+str(i)
            i4 = 'D'+str(i)
            i5 = 'E'+str(i)
            i6 = 'F'+str(i)
            i7 = 'G'+str(i)
            i8 = 'H'+str(i)
            f1 = '=Sheet1!B'+str(i)
            f2 = '=IF(OR(Sheet1!C'+str(i)+'="NULL",Sheet1!C'+str(i)+'=""),"",Sheet1!C'+str(i)+')'
            f3 = '=Sheet1!B'+str(i)
            f4 = '=IF(OR(Sheet1!L'+str(i)+'="NULL",Sheet1!L'+str(i)+'=""),"",Sheet1!L'+str(i)+')'
            # =IF(OR(Sheet1!L'+str(i)+'="NULL",Sheet1!L'+str(i)+'="",Sheet1!L'+str(i)+'="NO_ATTRIBUTE",Sheet1!L'+str(i)+'="null"),"",Sheet1!L'+str(i)+')'
            f5 = '=Sheet1!B'+str(i)
            f6 = '=IF(OR(Sheet1!M'+str(i)+'="NULL",Sheet1!M'+str(i)+'=""),"",Sheet1!M'+str(i)+')'
            # =IF(OR(Sheet1!M'+str(i)+'="NULL",Sheet1!M'+str(i)+'="",Sheet1!M'+str(i)+'="NO_ATTRIBUTE",Sheet1!M'+str(i)+'="null"),"",Sheet1!M'+str(i)+')'
            f7 = '=Sheet1!N'+str(i)
            f8 = '=Sheet1!O'+str(i)
            sheet_obj1[i1]=f1
            sheet_obj1[i2]=f2
            sheet_obj1[i3]=f3
            sheet_obj1[i4]=f4
            sheet_obj1[i5]=f5
            sheet_obj1[i6]=f6
            sheet_obj1[i7]=f7
            sheet_obj1[i8]=f8
        i1 = 'C'+str(n)
        i2 = 'G'+str(n)
        i3 = 'E'+str(n)
        i4 = 'H'+str(n)
        f1 = 'FILTER(Sheet1!D2:D'+str(m-1)+',Sheet1!P2:P'+str(m-1)+'="RAMP id not in AWS")'
        f2 = 'FILTER(Sheet1!P2:P'+str(m-1)+',Sheet1!P2:P'+str(m-1)+'="RAMP id not in AWS")'
        f3 = 'FILTER(Sheet1!H2:H'+str(p-1)+',Sheet1!Q2:Q'+str(p-1)+'="Predix id not in AWS")'
        f4 = 'FILTER(Sheet1!Q2:Q'+str(p-1)+',Sheet1!Q2:Q'+str(p-1)+'="Predix id not in AWS")'
        sheet_obj1[i1]=f1
        sheet_obj1[i2]=f2
        sheet_obj1[i3]=f3
        sheet_obj1[i4]=f4
        obj.save(path)
    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")