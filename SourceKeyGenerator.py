from matplotlib.pyplot import text
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles import Font

def SKG(file):
    try:
        print(Fore.RESET)
        path = './excel files/'+file
        
        df = pd.read_excel(path,sheet_name="Sheet1")
        n = df.count()[0]+2
        obj = openpyxl.load_workbook(path.strip())
        sheet_obj = obj["Sheet1"]
        f10 = openpyxl.styles.fills.PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
        sheet_obj['B1']='trim SourceKey'
        sheet_obj['C1']= 'Ready to use SourceKey'

        for y in range(1,3+1):
            # sheet_obj.cell(row=1,column=y).fill = f10
            sheet_obj.cell(row=1,column=y).font = Font(bold=True)

        for i in range(2,n):
            i1 = 'B'+str(i)
            i2 = 'C'+str(i)
            f1 = '=TRIM(A'+str(i)+')'
            f2 = '=IF(OR(ISNUMBER(SEARCH("GE_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("GE_OFW_",B'+str(i)+'))),REPLACE(B'+str(i)+',1,7,""),IF(OR(ISNUMBER(SEARCH("VES_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("CLP_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("ENR_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("ALS_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("NOR_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("TAC_ONW_",B'+str(i)+')),ISNUMBER(SEARCH("GAM_ONW_",B'+str(i)+'))),REPLACE(B'+str(i)+',1,8,""),TRIM(B'+str(i)+')))'
            sheet_obj[i1] = f1 
            sheet_obj[i2] = f2
        obj.save(path)
    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.")