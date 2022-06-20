import keyword
from matplotlib.pyplot import text
import xlsxwriter
from datetime import date;
from colorama import Fore
import openpyxl
import pandas as pd
from openpyxl.styles.fills import PatternFill
from openpyxl.styles import Font, colors
import subprocess
import os
import time
import numpy as np
from pynput.keyboard import Key,Controller
from pymysql import NULL
def MA(file1):
    try: 

        print(Fore.RESET)
        path = './excel files/'+file1
        print(path)

        df1 = pd.read_excel(path,sheet_name="AWS")
        df2 = pd.read_excel(path,sheet_name="RAMP")
        df3 = pd.read_excel(path,sheet_name="Predix")
        n = df1.count()[0] + 2
        m = df2.count()[0] + 3
        p = df3.count()[0] + 2
        today = date.today()
        obj = openpyxl.load_workbook(path.strip())
        
        ws1 = obj.create_sheet("Sheet2")
        ws1.title= "op_"+str(today)+"(AWS vs RAMP)"
        ws1 = obj.create_sheet("Sheet3")
        ws1.title= "op_"+str(today)+"(AWS vs Predix)"
        sheet_obj1 = obj["AWS"]
        sheet_obj2 = obj["RAMP"]
        sheet_obj3 = obj["op_"+str(today)+"(AWS vs RAMP)"]
        sheet_obj4 = obj["Predix"]
        sheet_obj5 = obj["op_"+str(today)+"(AWS vs Predix)"]
        
        sheet_obj1.insert_cols(0,amount=1)
        sheet_obj2.insert_cols(0,amount=1)
        sheet_obj4.insert_cols(0,amount=1)
        sheet_obj4.insert_cols(3,amount=1)

        #Naming of column
        sheet_obj1['A1']='TRIM _id'
        sheet_obj2['A1']='TRIM Turbine Serial Number'
        sheet_obj2['Q1']='ID in AWS?'
        sheet_obj2['R1']='name'
        sheet_obj3['A1']='Asset From AWS'
        sheet_obj3['B1']='Asset From RAMP'

        sheet_obj3['C1']='Frequency(AWS)'
        sheet_obj3['D1']='Frequency(RAMP)'
        sheet_obj3['E1']='Height(AWS)'
        sheet_obj3['F1']='Height(RAMP)'
        sheet_obj3['G1']='Air Desnsity(AWS)'
        sheet_obj3['H1']='Air Desnsity(RAMP)'

        sheet_obj3['I1']='Altitude(AWS)'
        sheet_obj3['J1']='Altitude(RAMP)'
        sheet_obj3['K1']='Make(AWS)'
        sheet_obj3['L1']='Make(RAMP)'
        sheet_obj3['M1']='Diameter(AWS)'
        sheet_obj3['N1']='Diameter(RAMP)'
        sheet_obj3['O1']='Short Name(AWS)'
        sheet_obj3['P1']='Short Name(RAMP)'
        sheet_obj3['Q1']='Rating(AWS)'
        sheet_obj3['R1']='R_Rating(RAMP)'
        sheet_obj3['S1']='Rating(RAMP)'
        sheet_obj3['T1']='COD(AWS)'
        sheet_obj3['U1']='R_COD(RAMP)'
        sheet_obj3['V1']='COD(RAMP)'
        sheet_obj3['W1']='model(AWS)'
        sheet_obj3['X1']='Current_Model(RAMP)'
        sheet_obj3['Y1']='Model(RAMP)'
        sheet_obj3['Z1']='name (AWS)'
        sheet_obj3['AA1']='name (RAMP)'
        sheet_obj3['AB1']='discrepancy Status'

        sheet_obj4['A1'] = 'trim SourceKey'
        sheet_obj4['C1'] = 'trim SerialNumber'
        sheet_obj4['N1'] = "id not in AWS?"
        sheet_obj5['A1']='Asset From AWS'
        sheet_obj5['B1']='Asset From Predix'

        sheet_obj5['C1']='Frequency(AWS)'
        sheet_obj5['D1']='Frequency(Predix)'
        sheet_obj5['E1']='Height(AWS)'
        sheet_obj5['F1']='Height(Predix)'
        sheet_obj5['G1']='Air Desnsity(AWS)'
        sheet_obj5['H1']='Air Desnsity(Predix)'

        sheet_obj5['I1']='Altitude(AWS)'
        sheet_obj5['J1']='Altitude(Predix)'
        sheet_obj5['K1']='Diameter(AWS)'
        sheet_obj5['L1']='Diameter(Predix)'
        sheet_obj5['M1']='Short Name(AWS)'
        sheet_obj5['N1']='Short Name(Predix)'
        sheet_obj5['O1']='Rating(AWS)'
        sheet_obj5['P1']='Rating(Predix)'
        sheet_obj5['Q1']='model(AWS)'
        sheet_obj5['R1']='Model(Predix)'
        sheet_obj5['S1']='name (AWS)'
        sheet_obj5['T1']='name (Predix)'
        sheet_obj5['U1']='discrepancy Status'

        print('n',n)
        for i in range(2,n):
            index1 = 'A'+str(i)
            index2 = 'A'+str(i)
            index3 = 'B'+str(i)
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=AWS!A'+str(i) 
            f3 = '=AWS!A'+str(i)
            sheet_obj1[index1] = f1
            sheet_obj3[index2] = f2
            sheet_obj3[index3] = f3
            sheet_obj5[index2] = f2
            sheet_obj5[index3] = f3

        print('m',m) 
        for i in range(2,m):

            index1 = 'A'+str(i)
            index2 = 'Q'+str(i)
            index3 = 'R'+str(i)
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=if(ISNA(VLOOKUP(A'+str(i)+',AWS!A:B,1,FALSE)),"RAMP id not in AWS",VLOOKUP(A'+str(i)+',AWS!A:B,1,FALSE))'
            f3 = '=P'+str(i)+'&" "&I'+str(i)+''
            sheet_obj2[index1] = f1
            sheet_obj2[index2] = f2
            sheet_obj2[index3]=f3
        
        print('p',p)
        for i in range(2,p):
            index1 = 'A'+str(i)
            index2 = 'C'+str(i)
            index3 = 'N'+str(i)
            f1 = '=TRIM(B'+str(i)+')'
            f2 = '=TRIM(D'+str(i)+')'
            f3 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',AWS!A:A,1,FALSE)),ISNA(VLOOKUP(C'+str(i)+',AWS!A:A,1,FALSE))),"Predix id not in AWS",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',AWS!A:A,1,FALSE))),VLOOKUP(A'+str(i)+',AWS!A:A,1,FALSE),IF(NOT(ISNA(VLOOKUP(C'+str(i)+',AWS!A:A,1,FALSE))),VLOOKUP(C'+str(i)+',AWS!A:A,1,FALSE),"Predix id not in AWS")))'
            sheet_obj4[index1] = f1
            sheet_obj4[index2] = f2
            sheet_obj4[index3] = f3
        
        for i in range(2,n):
            i1 = 'C'+str(i)
            i2 = 'E'+str(i)
            i3 = 'G'+str(i)
            i4 = 'I'+str(i)
            i5 = 'K'+str(i)
            i6 = 'M'+str(i)
            i7 = 'O'+str(i)
            i8 = 'Q'+str(i)
            i9 = 'T'+str(i)
            i10 = 'W'+str(i)
            i11 = 'Z'+str(i)
            i12 = 'S'+str(i)
            
            f1 = '=IF(AWS!C'+str(i)+'="NULL","",AWS!C'+str(i)+')'
            f2 = '=IF(AWS!D'+str(i)+'="NULL","",AWS!D'+str(i)+')'
            f3 = '=IF(AWS!E'+str(i)+'="NULL","",AWS!E'+str(i)+')'
            f4 = '=IF(AWS!F'+str(i)+'="NULL","",AWS!F'+str(i)+')'
            f5 = '=IF(AWS!G'+str(i)+'="NULL","",AWS!G'+str(i)+')'
            f6 = '=IF(AWS!H'+str(i)+'="NULL","",AWS!H'+str(i)+')'
            f7 = '=IF(trim(AWS!I'+str(i)+')="NULL","",trim(AWS!I'+str(i)+'))'
            f8 = '=IF(trim(AWS!J'+str(i)+')="NULL","",trim(AWS!J'+str(i)+'))'
            f9 = '=IF(AWS!K'+str(i)+'="NULL","",text(AWS!K'+str(i)+',"dd/mm/yyyy"))'
            f10 = '=IF(AWS!L'+str(i)+'="NULL","",AWS!L'+str(i)+')'
            f11 = '=IF(AWS!M'+str(i)+'="NULL","",lower(AWS!M'+str(i)+'))'

            sheet_obj3[i1]=f1
            sheet_obj3[i2]=f2
            sheet_obj3[i3]=f3
            sheet_obj3[i4]=f4
            sheet_obj3[i5]=f5
            sheet_obj3[i6]=f6
            sheet_obj3[i7]=f7
            sheet_obj3[i8]=f8
            sheet_obj3[i9]=f9
            sheet_obj3[i10]=f10
            sheet_obj3[i11]=f11

            sheet_obj5[i1] = f1
            sheet_obj5[i2] = f2
            sheet_obj5[i3] = f3
            sheet_obj5[i4] = f4
            sheet_obj5[i5] = f6
            sheet_obj5[i6] = f7
            sheet_obj5[i7] = f8
            sheet_obj5[i8] = f10
            sheet_obj5[i12] = f11
        
        for i in range(2,n):
            i1 = 'D'+str(i)
            i2 = 'F'+str(i)
            i3 = 'H'+str(i)
            i4 = 'J'+str(i) 
            i5 = 'L'+str(i)
            i6 = 'N'+str(i)
            i7 = 'P'+str(i)
            i8 = 'R'+str(i)
            i9 = 'S'+str(i)
            i10 = 'U'+str(i)
            i11 = 'V'+str(i)
            i12 = 'X'+str(i)
            i13 = 'Y'+str(i)
            i14 = 'AA'+str(i)
            
            f1 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:C,3,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:C,3,FALSE))=0,"",VLOOKUP(A'+str(i)+',RAMP!A:C,3,FALSE)))' #frequency
            f2 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:D,4,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:D,4,FALSE))=0,"",Round(VLOOKUP(A'+str(i)+',RAMP!A:D,4,FALSE),6) ))' #VLOOKUP(A1,RAMP!A:D,4,FALSE) #height
            f3 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:E,5,FALSE)),"AWS id not in RAMP",if(LEN(VLOOKUP(A'+str(i)+',RAMP!A:E,5,FALSE))=0,"",VLOOKUP(A'+str(i)+',RAMP!A:E,5,FALSE)))'#air density
            f4 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:F,6,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:F,6,FALSE))=0,"",VLOOKUP(A'+str(i)+',RAMP!A:F,6,FALSE)))'#altitude
            f5 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:G,7,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:G,7,FALSE))=0,"",VLOOKUP(A'+str(i)+',RAMP!A:G,7,FALSE)))'#make
            f6 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:H,8,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:H,8,FALSE))=0,"",Round(VLOOKUP(A'+str(i)+',RAMP!A:H,8,FALSE),6) ))' #diameter
            f7 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:I,9,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:I,9,FALSE))=0,"",trim(VLOOKUP(A'+str(i)+',RAMP!A:I,9,FALSE))))'#short name
            f8 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:J,10,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:J,10,FALSE))=0,"",trim(VLOOKUP(A'+str(i)+',RAMP!A:J,10,FALSE))))' #R_rating
            f9 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:K,11,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:K,11,FALSE))=0,"",trim(VLOOKUP(A'+str(i)+',RAMP!A:K,11,FALSE))))' #rating
            f10 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:L,12,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:L,12,FALSE))=0,"",text(DATE(YEAR(VLOOKUP(A'+str(i)+',RAMP!A:L,12,FALSE)),MONTH(VLOOKUP(A'+str(i)+',RAMP!A:L,12,FALSE)),DAY(VLOOKUP(A'+str(i)+',RAMP!A:L,12,FALSE))),"m/dd/yyyy") ))'#R_cod
            f11 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:M,13,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:M,13,FALSE))=0,"",text(DATE(YEAR(VLOOKUP(A'+str(i)+',RAMP!A:M,13,FALSE)),MONTH(VLOOKUP(A'+str(i)+',RAMP!A:M,13,FALSE)),DAY(VLOOKUP(A'+str(i)+',RAMP!A:M,13,FALSE))),"m/dd/yyyy") ))'#cod
            f12 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:N,14,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:N,14,FALSE))=0,"",VLOOKUP(A'+str(i)+',RAMP!A:N,14,FALSE)))'#r_model
            f13 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:O,15,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:O,15,FALSE))=0,"",VLOOKUP(A'+str(i)+',RAMP!A:O,15,FALSE)))'#model
            f14 = '=if(ISNA(VLOOKUP(A'+str(i)+',RAMP!A:R,18,FALSE)),"AWS id not in RAMP",if(len(VLOOKUP(A'+str(i)+',RAMP!A:R,18,FALSE))=0,"",lower(VLOOKUP(A'+str(i)+',RAMP!A:R,18,FALSE))))' #name
            sheet_obj3[i1]=f1
            sheet_obj3[i2]=f2
            sheet_obj3[i3]=f3
            sheet_obj3[i4]=f4
            sheet_obj3[i5]=f5
            sheet_obj3[i6]=f6
            sheet_obj3[i7]=f7
            sheet_obj3[i8]=f8
            sheet_obj3[i9]=f9
            sheet_obj3[i10]=f10
            sheet_obj3[i11]=f11
            sheet_obj3[i12]=f12
            sheet_obj3[i13]=f13
            sheet_obj3[i14]=f14
        
        for i in range(2,n):
            i1 = 'D'+str(i)
            i2 = 'F'+str(i)
            i3 = 'H'+str(i)
            i4 = 'J'+str(i)
            i5 = 'L'+str(i)
            i6 = 'N'+str(i)
            i7 = 'P'+str(i)
            i8 = 'R'+str(i)
            i9 = 'T'+str(i)

            f1 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:E,5,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:E,3,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:E,5,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:E,5,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:E,5,FALSE)="null"),"",ROUND(VLOOKUP(A'+str(i)+',Predix!A:E,5,FALSE),6)),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:E,3,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:E,3,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:E,3,FALSE)="null"),"",ROUND(VLOOKUP(A'+str(i)+',Predix!C:E,3,FALSE),6)),"AWS id not in Predix")))' #Frequency
            f2 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:F,6,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:F,4,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:F,6,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:F,6,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:F,6,FALSE)="null"),"",ROUND(VLOOKUP(A'+str(i)+',Predix!A:F,6,FALSE),6)),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:F,4,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:F,4,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:F,4,FALSE)="null"),"",ROUND(VLOOKUP(A'+str(i)+',Predix!C:F,4,FALSE),6)),"AWS id not in Predix")))' #hub height
            f3 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:G,7,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:G,5,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:G,7,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:G,7,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:G,7,FALSE)="NO_ATTRIBUTE",VLOOKUP(A'+str(i)+',Predix!A:G,7,FALSE)="null"),"",VALUE(VLOOKUP(A'+str(i)+',Predix!A:G,7,FALSE))),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:G,5,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:G,5,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:G,5,FALSE)="NO_ATTRIBUTE",VLOOKUP(A'+str(i)+',Predix!C:G,5,FALSE)="null"),"",VALUE(VLOOKUP(A'+str(i)+',Predix!C:G,5,FALSE))),"AWS id not in Predix")))' #air density
            f4 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:H,8,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:H,6,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:H,8,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:H,8,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:H,8,FALSE)="NA"),"",VALUE(VLOOKUP(A'+str(i)+',Predix!A:H,8,FALSE))),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:H,6,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:H,6,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:H,6,FALSE)="NA"),"",VALUE(VLOOKUP(A'+str(i)+',Predix!C:H,6,FALSE))),"AWS id not Predix")))' #altitude
            f5 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:I,9,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:I,7,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:I,9,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:I,9,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:I,9,FALSE)="NULL"),"",ROUND(VLOOKUP(A'+str(i)+',Predix!A:I,9,FALSE),6)),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:I,7,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:I,7,FALSE)="NULL"),"",ROUND(VLOOKUP(A'+str(i)+',Predix!C:I,7,FALSE),6))))'#rotor diameter
            f6 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE)="null",VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE)="NULL",VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE)=""),"",VLOOKUP(A'+str(i)+',Predix!A:J,10,FALSE)),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE)="null",VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE)="NULL",VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE)=""),"",VLOOKUP(A'+str(i)+',Predix!C:J,8,FALSE)))))'#Shortname
            f7 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:K,11,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:K,9,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:K,11,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:K,11,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!A:K,11,FALSE)="NULL"),"",TRIM(VLOOKUP(A'+str(i)+',Predix!A:K,11,FALSE))),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:K,9,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:K,9,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:K,9,FALSE)="NULL"),"",VLOOKUP(A'+str(i)+',Predix!C:K,9,FALSE)),TRIM(VLOOKUP(A'+str(i)+',Predix!C:K,9,FALSE)))))'#rating
            f8 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:L,12,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:L,10,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:L,12,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!A:L,12,FALSE))=0,TRIM(VLOOKUP(A'+str(i)+',Predix!A:L,12,FALSE))="null"),"",VLOOKUP(A'+str(i)+',Predix!A:L,12,FALSE)),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:L,10,FALSE))),IF(OR(LEN(VLOOKUP(A'+str(i)+',Predix!C:L,10,FALSE))=0,VLOOKUP(A'+str(i)+',Predix!C:L,10,FALSE)="null"),"",VLOOKUP(A'+str(i)+',Predix!C:L,10,FALSE)),"AWS id not in Predix")))'#model
            f9 = '=IF(AND(ISNA(VLOOKUP(A'+str(i)+',Predix!A:M,13,FALSE)),ISNA(VLOOKUP(A'+str(i)+',Predix!C:M,11,FALSE))),"AWS id not in Predix",IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!A:M,13,FALSE))),IF(LEN(VLOOKUP(A'+str(i)+',Predix!A:M,13,FALSE))=0,"",LOWER(VLOOKUP(A'+str(i)+',Predix!A:M,13,FALSE))),IF(NOT(ISNA(VLOOKUP(A'+str(i)+',Predix!C:M,11,FALSE))),IF(LEN(VLOOKUP(A'+str(i)+',Predix!C:M,11,FALSE))=0,"",LOWER(VLOOKUP(A'+str(i)+',Predix!C:M,11,FALSE))),"AWS id not in Predix")))'#name

            sheet_obj5[i1]=f1
            sheet_obj5[i2]=f2
            sheet_obj5[i3]=f3
            sheet_obj5[i4]=f4
            sheet_obj5[i5]=f5
            sheet_obj5[i6]=f6
            sheet_obj5[i7]=f7
            sheet_obj5[i8]=f8
            sheet_obj5[i9]=f9
        
        obj.save(path)
        
        print("opening...")
        excel_prog =r'C:\Program Files (x86)\Microsoft Office\root\Office16\EXCEL.EXE'
        time.sleep(5)
        subprocess.Popen([excel_prog, os.path.join(os.getcwd(), path)])

        time.sleep(100)
        keyboard = Controller()
        with keyboard.pressed(Key.ctrl):
            keyboard.press('s')
            keyboard.release('s')
        print("Saving...")
        time.sleep(25)
        with keyboard.pressed(Key.alt):
            keyboard.press(Key.f4)
            keyboard.release(Key.f4)
        print("closing...")
        time.sleep(10)
        obj = openpyxl.load_workbook(path, data_only=True)
        sheet_obj3 = obj["op_"+str(today)+"(AWS vs RAMP)"]
        sheet_obj5 = obj["op_"+str(today)+"(AWS vs Predix)"]

        
        # not matching f1
        # id not in RAMP or Predix f2
        # id not in AWS f3
        # values matching f4

        f1 = openpyxl.styles.fills.PatternFill(start_color='FCE4D6',end_color='FCE4D6',fill_type='solid')
        f2 = openpyxl.styles.fills.PatternFill(start_color='D9E1F2',end_color='D9E1F2',fill_type='solid')
        f3 = openpyxl.styles.fills.PatternFill(start_color='FFF2CC',end_color='FFF2CC',fill_type='solid')
        f4 = openpyxl.styles.fills.PatternFill(start_color='E2EFDA',end_color='E2EFDA',fill_type='solid')
        
        print("Filling color.. AWS vs RAMP")
        for i in range(2,n):
            
            index1 = 'C'+str(i)
            index2 = 'D'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            # 
            index1 = 'E'+str(i)
            index2 = 'F'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            # 
            index1 = 'G'+str(i)
            index2 = 'H'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            # 
            index1 = 'I'+str(i)
            index2 = 'J'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):

                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            # 
            index1 = 'K'+str(i)
            index2 = 'L'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            # 
            index1 = 'M'+str(i)
            index2 = 'N'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            # 
            index1 = 'O'+str(i)
            index2 = 'P'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            #repowered attributes
            index1 = 'Q'+str(i)
            index2 = 'R'+str(i)
            index3 = 'S'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
                sheet_obj3[index3].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                    # 
                    if sheet_obj3[index1].value != sheet_obj3[index3].value:
                        if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index3].value==NULL or sheet_obj3[index3].value==None)):
                            sheet_obj3[index1].fill = f1
                            sheet_obj3[index3].fill = f1
                        else:
                            sheet_obj3[index1]=""
                            sheet_obj3[index2]=""
                            sheet_obj3[index1].fill = f4
                            sheet_obj3[index3].fill = f4
                    elif sheet_obj3[index1].value == sheet_obj3[index3].value:
                        sheet_obj3[index1].fill = f4
                        sheet_obj3[index3].fill = f4
                    # 
                else: 
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            index1 = 'T'+str(i)
            index2 = 'U'+str(i)
            index3 = 'V'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
                sheet_obj3[index3].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                    # 
                    if sheet_obj3[index1].value != sheet_obj3[index3].value:
                        if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index3].value==NULL or sheet_obj3[index3].value==None)):
                            sheet_obj3[index1].fill = f1
                            sheet_obj3[index3].fill = f1
                        else:
                            sheet_obj3[index1]=""
                            sheet_obj3[index2]=""
                            sheet_obj3[index1].fill = f4
                            sheet_obj3[index3].fill = f4
                    elif sheet_obj3[index1].value == sheet_obj3[index3].value:
                        sheet_obj3[index1].fill = f4
                        sheet_obj3[index3].fill = f4
                    # 
                else: 
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            index1 = 'W'+str(i)
            index2 = 'X'+str(i)
            index3 = 'Y'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
                sheet_obj3[index3].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                    # 
                    if sheet_obj3[index1].value != sheet_obj3[index3].value:
                        if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index3].value==NULL or sheet_obj3[index3].value==None)):
                            sheet_obj3[index1].fill = f1
                            sheet_obj3[index3].fill = f1
                        else:
                            sheet_obj3[index1] = ""
                            sheet_obj3[index2] = ""
                            sheet_obj3[index1].fill = f4
                            sheet_obj3[index3].fill = f4
                    elif sheet_obj3[index1].value == sheet_obj3[index3].value:
                        sheet_obj3[index1].fill = f4
                        sheet_obj3[index3].fill = f4
                    # 
                else: 
                    sheet_obj3[index1]=""
                    sheet_obj3[index2]=""
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
            index1 = 'Z'+str(i)
            index2 = 'AA'+str(i)
            if sheet_obj3[index2].value=="AWS id not in RAMP":
                if (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None):
                    sheet_obj3[index1]=None
                sheet_obj3[index1].fill = f2
                sheet_obj3[index2].fill = f2
            elif sheet_obj3[index1].value != sheet_obj3[index2].value:
                if not( (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None) and (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None)):
                    if (sheet_obj3[index1].value==NULL or sheet_obj3[index1].value==None):
                        sheet_obj1[index1]=None
                    if (sheet_obj3[index2].value==NULL or sheet_obj3[index2].value==None):
                        sheet_obj3[index2]=None
                    sheet_obj3[index1].fill = f1
                    sheet_obj3[index2].fill = f1
                else:
                    sheet_obj3[index1]=None
                    sheet_obj3[index2]=None
                    sheet_obj3[index1].fill = f4
                    sheet_obj3[index2].fill = f4
            
            elif sheet_obj3[index1].value == sheet_obj3[index2].value:
                sheet_obj3[index1].fill = f4
                sheet_obj3[index2].fill = f4
        
        print("Filling color.. AWS vs Predix")
        for i in range(2,n):
            
            index1 = 'C'+str(i)
            index2 = 'D'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4          
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'E'+str(i)
            index2 = 'F'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4            
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'G'+str(i)
            index2 = 'H'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4           
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'I'+str(i)
            index2 = 'J'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):

                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4            
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'K'+str(i)
            index2 = 'L'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4           
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'M'+str(i)
            index2 = 'N'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4            
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'O'+str(i)
            index2 = 'P'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4            
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'Q'+str(i)
            index2 = 'R'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
            # 
            index1 = 'S'+str(i)
            index2 = 'T'+str(i)
            if sheet_obj5[index2].value=="AWS id not in Predix":
                
                sheet_obj5[index1].fill = f2
                sheet_obj5[index2].fill = f2
            elif sheet_obj5[index1].value != sheet_obj5[index2].value:
                if not( (sheet_obj5[index1].value==NULL or sheet_obj5[index1].value==None) and (sheet_obj5[index2].value==NULL or sheet_obj5[index2].value==None)):
                    
                    sheet_obj5[index1].fill = f1
                    sheet_obj5[index2].fill = f1
                else:
                    sheet_obj5[index1]=""
                    sheet_obj5[index2]=""
                    sheet_obj5[index1].fill = f4
                    sheet_obj5[index2].fill = f4          
            elif sheet_obj5[index1].value == sheet_obj5[index2].value:
                sheet_obj5[index1].fill = f4
                sheet_obj5[index2].fill = f4
        
        obj.save(path)

        
        obj = openpyxl.load_workbook(path.strip())
        sheet_obj3 = obj["op_"+str(today)+"(AWS vs RAMP)"]
        sheet_obj5 = obj["op_"+str(today)+"(AWS vs Predix)"]
         

        df2 = pd.read_excel(path,sheet_name="RAMP") 
        df3 = pd.read_excel(path,sheet_name="Predix")
        b = df2.iloc[:,16]
        d = df3.iloc[:,13]
        l = list(np.where(b == "RAMP id not in AWS")[0])
        r = list(np.where(d == "Predix id not in AWS")[0])
        L = len(l)
        R = len(r)

        # j = 46787
        j = n
        # print('j=',j)
        # print("h1")
        # print("L=",L)
        for i in range(0,L):
            index1 = 'B'+str(j)
            index2 = 'C'+str(j)
            index3 = 'E'+str(j)
            index4 = 'G'+str(j)
            index5 = 'I'+str(j)
            index6 = 'K'+str(j)
            index7 = 'M'+str(j)
            index8 = 'O'+str(j)
            index9 = 'Q'+str(j)
            index10 = 'T'+str(j)
            index11 = 'W'+str(j)
            index12 = 'Z'+str(j)

            i1 = 'D'+str(j)
            i2 = 'F'+str(j)
            i3 = 'H'+str(j)
            i4 = 'J'+str(j)
            i5 = 'L'+str(j)
            i6 = 'N'+str(j)
            i7 = 'P'+str(j)
            i8 = 'R'+str(j)
            i9 = 'S'+str(j)
            i10 = 'U'+str(j)
            i11 = 'V'+str(j)
            i12 = 'X'+str(j)
            i13 = 'Y'+str(j)
            i14 = 'AA'+str(j)
            
            sheet_obj3[index1].value = df2[df2.columns[1]][l[i]]
            sheet_obj3[index2]="RAMP id not in AWS"
            sheet_obj3[index2].fill = f3
            # sheet_obj3[i1]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:C,3,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:C,3,FALSE))' #frequency
            sheet_obj3[index3]="RAMP id not in AWS"
            sheet_obj3[index3].fill = f3
            # sheet_obj3[i2]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:D,4,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:D,4,FALSE))' #Tower height
            sheet_obj3[index4]="RAMP id not in AWS"
            sheet_obj3[index4].fill = f3
            # sheet_obj3[i3]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:E,5,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:E,5,FALSE))' #air density
            sheet_obj3[index5]="RAMP id not in AWS"
            sheet_obj3[index5].fill = f3
            # sheet_obj3[i4]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:F,6,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:F,6,FALSE))' #hub 
            sheet_obj3[index6]="RAMP id not in AWS"
            sheet_obj3[index6].fill = f3
            # sheet_obj3[i5]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:G,7,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:G,7,FALSE))' #make
            sheet_obj3[index7]="RAMP id not in AWS"
            sheet_obj3[index7].fill = f3
            # sheet_obj3[i6]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:H,8,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:H,8,FALSE))' #rotor diameter
            sheet_obj3[index8]="RAMP id not in AWS"
            sheet_obj3[index8].fill = f3
            # sheet_obj3[i7]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:I,9,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:I,9,FALSE))' #pad number/short name
            sheet_obj3[index9]="RAMP id not in AWS"
            sheet_obj3[index9].fill = f3
            # sheet_obj3[i8]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:J,10,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:J,10,FALSE))' #r_new_rating
            # sheet_obj3[i9]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:K,11,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:K,11,FALSE))' #rating
            sheet_obj3[index10]="RAMP id not in AWS"
            sheet_obj3[index10].fill = f3
            # sheet_obj3[i10]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:L,12,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:L,12,FALSE))' #r_cod
            # sheet_obj3[i11]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:M,13,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:M,13,FALSE))' #cod
            sheet_obj3[index11]="RAMP id not in AWS"
            sheet_obj3[index11].fill = f3
            # sheet_obj3[i12]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:N,14,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:N,14,FALSE))' #current model
            # sheet_obj3[i13]='=if(len(VLOOKUP(B'+str(j)+',RAMP!A:O,15,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:O,15,FALSE))' #model
            sheet_obj3[index12]="RAMP id not in AWS"
            sheet_obj3[index12].fill = f3
            # sheet_obj3[i14]='=IF(LEN(VLOOKUP(B'+str(j)+',RAMP!A:R,18,FALSE))=0,"",VLOOKUP(B'+str(j)+',RAMP!A:R,18,FALSE))' #properties.name
            j=j+1

        j = n
        for i in range(0,R):
            index1 = 'B'+str(j)
            index2 = 'C'+str(j)
            index3 = 'E'+str(j)
            index4 = 'G'+str(j)
            index5 = 'I'+str(j)
            index6 = 'K'+str(j)
            index7 = 'M'+str(j)
            index8 = 'O'+str(j)
            index9 = 'Q'+str(j)
            index10 = 'S'+str(j)
            # index11 = 'W'+str(j)
            # index12 = 'Z'+str(j)

            # i1 = 'D'+str(j)
            # i2 = 'F'+str(j)
            # i3 = 'H'+str(j)
            # i4 = 'J'+str(j)
            # i5 = 'L'+str(j)
            # i6 = 'N'+str(j)
            # i7 = 'P'+str(j)
            # i8 = 'R'+str(j)
            # i9 = 'S'+str(j)
            # i10 = 'U'+str(j)
            # i11 = 'V'+str(j)
            # i12 = 'X'+str(j)
            # i13 = 'Y'+str(j)
            # i14 = 'AA'+str(j)
            
            sheet_obj5[index1].value = df3[df3.columns[1]][r[i]]
            sheet_obj5[index2]="Predix id not in AWS"
            sheet_obj5[index2].fill = f3
            # sheet_obj5[i1]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:C,3,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:C,3,FALSE))' #frequency
            sheet_obj5[index3]="Predix id not in AWS"
            sheet_obj5[index3].fill = f3
            # sheet_obj5[i2]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:D,4,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:D,4,FALSE))' #Tower height
            sheet_obj5[index4]="Predix id not in AWS"
            sheet_obj5[index4].fill = f3
            # sheet_obj5[i3]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:E,5,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:E,5,FALSE))' #air density
            sheet_obj5[index5]="Predix id not in AWS"
            sheet_obj5[index5].fill = f3
            # sheet_obj5[i4]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:F,6,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:F,6,FALSE))' #hub 
            sheet_obj5[index6]="Predix id not in AWS"
            sheet_obj5[index6].fill = f3
            # sheet_obj5[i5]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:G,7,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:G,7,FALSE))' #make
            sheet_obj5[index7]="Predix id not in AWS"
            sheet_obj5[index7].fill = f3
            # sheet_obj5[i6]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:H,8,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:H,8,FALSE))' #rotor diameter
            sheet_obj5[index8]="Predix id not in AWS"
            sheet_obj5[index8].fill = f3
            # sheet_obj5[i7]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:I,9,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:I,9,FALSE))' #pad number/short name
            sheet_obj5[index9]="Predix id not in AWS"
            sheet_obj5[index9].fill = f3
            # sheet_obj5[i8]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:J,10,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:J,10,FALSE))' #r_new_rating
            # sheet_obj5[i9]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:K,11,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:K,11,FALSE))' #rating
            sheet_obj5[index10]="Predix id not in AWS"
            sheet_obj5[index10].fill = f3
            # sheet_obj5[i10]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:L,12,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:L,12,FALSE))' #r_cod
            # sheet_obj5[i11]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:M,13,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:M,13,FALSE))' #cod
            sheet_obj5[index11]="Predix id not in AWS"
            sheet_obj5[index11].fill = f3
            # sheet_obj5[i12]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:N,14,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:N,14,FALSE))' #current model
            # sheet_obj5[i13]='=if(len(VLOOKUP(B'+str(j)+',Predix!A:O,15,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:O,15,FALSE))' #model
            sheet_obj5[index12]="Predix id not in AWS"
            sheet_obj5[index12].fill = f3
            # sheet_obj5[i14]='=IF(LEN(VLOOKUP(B'+str(j)+',Predix!A:R,18,FALSE))=0,"",VLOOKUP(B'+str(j)+',Predix!A:R,18,FALSE))' #properties.name
            j=j+1
        
        for i in range(2,n):
            i1 = 'AB'+str(i)
            i2 = 'U'+str(i)
            f1 = '=IF(AND(C'+str(i)+'=D'+str(i)+',E'+str(i)+'=F'+str(i)+',G'+str(i)+'=H'+str(i)+',I'+str(i)+'=J'+str(i)+',K'+str(i)+'=L'+str(i)+',M'+str(i)+'=N'+str(i)+',O'+str(i)+'=P'+str(i)+',OR(Q'+str(i)+'=R'+str(i)+',Q'+str(i)+'=S'+str(i)+'),OR(T'+str(i)+'=U'+str(i)+',T'+str(i)+'=V'+str(i)+'),OR(W'+str(i)+'=X'+str(i)+',W'+str(i)+'=Y'+str(i)+'),Z'+str(i)+'=AA'+str(i)+'),"matching",IF(D'+str(i)+'="AWS id not in RAMP","AWS id not in RAMP","not matching"))'
            f2 = '=IF(AND(C'+str(i)+'=D'+str(i)+',E'+str(i)+'=F'+str(i)+',G'+str(i)+'=H'+str(i)+',I'+str(i)+'=J'+str(i)+',K'+str(i)+'=L'+str(i)+',M'+str(i)+'=N'+str(i)+',O'+str(i)+'=P'+str(i)+',Q'+str(i)+'=R'+str(i)+',S'+str(i)+'=T'+str(i)+'),"matching",IF(D'+str(i)+'="AWS id not in Predix","AWS id not in Predix","not matching"))'
            sheet_obj3[i1]=f1
            sheet_obj5[i2]=f2
        print("Final Saving...")
        obj.save(path)

    except Exception as e:
        print(e)
        print (Fore.RED + "Error : The file does not found")
        return ("An Error has occured, pls verify")
    print(Fore.GREEN + "###################### Successfully the excel file has been read/written. ##############################")
    return("Successfully the excel file has been read/written.") 